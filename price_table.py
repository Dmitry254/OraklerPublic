import openpyxl


def fill_table_wizards(wizards):
    wb = open_table("history_orakler")
    sheet_name = 'prices'
    recreate_sheet(wb, sheet_name)
    sheet = wb[sheet_name]
    row = sheet.max_row
    if row == 1:
        row -= 2
    try:
        for level in range(0, 6):
            list_name = 'Level ' + str(level)
            if list_name in wizards.keys():
                row += 2
                sheet.cell(row=row, column=1).value = f"Level {level}"
                prices_list = wizards[list_name]
                row += 1
                for price_number in range(len(prices_list)):
                    sheet.cell(row=row, column=price_number+1).value = prices_list[price_number]
                    if price_number == 0:
                        sheet.cell(row=level, column=15).value = list_name
                        sheet.cell(row=level, column=16).value = prices_list[price_number]
            else:
                continue
    finally:
        close_table(wb, "history_orakler")


def fill_table_scepters(scepters):
    wb = open_table("history_orakler")
    sheet_name = 'prices'
    sheet = wb[sheet_name]
    row = sheet.max_row
    if row == 1:
        row -= 2
    try:
        for scepter_rc in range(0, 45):
            list_name = f"RC {scepter_rc}0-{scepter_rc}9"
            if list_name in scepters.keys():
                row += 2
                sheet.cell(row=row, column=1).value = f"RC {scepter_rc}0-{scepter_rc+1}9"
                prices_list = scepters[list_name]
                row += 1
                for price_number in range(len(prices_list)):
                    sheet.cell(row=row, column=price_number+1).value = prices_list[price_number]
                    if price_number == 0:
                        sheet.cell(row=scepter_rc, column=18).value = list_name
                        sheet.cell(row=scepter_rc, column=19).value = prices_list[price_number]
    finally:
        close_table(wb, "history_orakler")



def fill_table_robi(robies):
    wb = open_table("history")
    sheet_name = 'prices'
    sheet = wb[sheet_name]
    row = sheet.max_row
    if row == 1:
        row -= 2
    try:
        for level in range(0, 10):
            list_name = 'robi' + str(level)
            if list_name in robies.keys():
                row += 2
                sheet.cell(row=row, column=1).value = f"Robi {level}"
                robi_list = robies[list_name]
                row += 1
                for robi_number in range(len(robi_list)):
                    sheet.cell(row=row, column=robi_number+1).value = robi_list[robi_number]
            else:
                continue
    finally:
        close_table(wb, "history_orakler")


def open_table(table_name):
    wb = openpyxl.load_workbook(f"{table_name}.xlsx")
    return wb


def close_table(wb, table_name):
    wb.save(f'{table_name}.xlsx')


def recreate_sheet(wb, sheet_name):
    sheet = wb[sheet_name]
    wb.remove(sheet)
    wb.create_sheet(sheet_name)


def fill_energy_value():
    wb = open_table()
    sheet = wb['price']
    for level in range(0, 40):
        sheet.cell(row=3+level, column=5).value = f"Energy {level}00-{level + 1}00"
    close_table(wb)
